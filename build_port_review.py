# -*- coding: utf-8 -*-
"""
Build Port Review DataFrame from local files.
Replaces the legacy Port Review pipeline.

Produces the same 26-column output as the existing Port_Review_YYYYMMDD.xlsx.

Cash columns are stubbed (set to 0) pending final source confirmation.

@author: auto-generated
"""
from __future__ import annotations

import os
import re
import glob
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import snowflake.connector
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization


# ===========================================================================
# CONFIGURATION
# ===========================================================================

# Snowflake (for index constituents only)
SF_KEY_PATH = r"X:\PM & Operations\Private Key\OPS_TEAM_private_key.p8"
SF_USER = "OPS_TEAM"
SF_ACCOUNT = "uib64437.us-east-1"
SF_WAREHOUSE = "COMPUTE_WH"
SF_DATABASE = "PROD"
SF_SCHEMA = "PUBLIC"

# Local file paths
BASE_DIR = Path(r"X:\PM & Operations\Portfolio Management\Daily Files")
RAW_DATA_DIR = BASE_DIR / "000 Raw Data" / "Basket & Position Files"
DIVIDEND_DIR = BASE_DIR / "Dividend Schedule"
BNY_CASH_DIR = BASE_DIR / "BNY Cash Balance"
BBH_CASH_DIR = BASE_DIR / "BBH Cash Balance"
BNY_CUSTODY_DIR = BASE_DIR / "BNY Custody"
BBH_CUSTODY_DIR = BASE_DIR / "BBH Custody"

MAP_FUND_PATH = Path(r"X:\PM & Operations\Projects\py\data_fields\custom\mapFUND.csv")

# Output columns in exact order (matches template)
OUTPUT_COLUMNS = [
    "Fund Ticker",
    "Attribute",
    "Assets",
    "Date",
    "Error Check",
    "Active Weights (Abs)",
    "Custody Cash (USD) BNY",
    "Custody Cash (USD) BBH",
    "Foreign Ccy BNY",
    "Foreign Ccy BBH",
    "Foreign Ccy",
    "CFC Cash",
    "Custody Cash (USD)",
    "Pending (USD) BNY",
    "Pending CIL",
    "Dist Pay",
    "Custody Cash (USD) - Adj",
    "Actual Cash",
    "Accrued Cash",
    "Net Cash",
    "Net Cash (Futures Adj)",
    "Rank 1",
    "Rank 2",
    "Rank 3",
    "Reinvest Flag",
    "Raise Flag",
    "Comment",
]

ATTRIBUTE_SORT_ORDER = {
    "THEMATIC": 0,
    "OPTIONS": 1,
    "INTERNATIONAL": 2,
    "INCOME": 3,
    "FI / DERIVS": 4,
    "FI/DERIVS": 4,
    "COMMODITIES": 5,
}

REINVEST_CUSTODY_CASH_ADJ_THRESHOLD = 0.0010  # 10 bps
REINVEST_NET_CASH_FUTURES_ADJ_THRESHOLD = 0.0015  # 15 bps
RAISE_CUSTODY_CASH_ADJ_THRESHOLD = -0.0004  # -4 bps


# ===========================================================================
# HELPERS
# ===========================================================================

def _sf_connect():
    """Connect to Snowflake using key-pair auth."""
    with open(SF_KEY_PATH, "rb") as f:
        key_data = f.read()
    pk = serialization.load_pem_private_key(key_data, password=None, backend=default_backend())
    pkb = pk.private_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    )
    return snowflake.connector.connect(
        user=SF_USER,
        account=SF_ACCOUNT,
        private_key=pkb,
        warehouse=SF_WAREHOUSE,
        database=SF_DATABASE,
        schema=SF_SCHEMA,
        role="",
        ocsp_fail_open=True,
    )


def _find_sei_folder(report_date: date) -> Path:
    """
    Find the SEI dated folder for the given report date.
    Tries the exact date, then the next business day, then falls back to
    the most recent available SEI folder.
    """
    folder_name = f"SEI {report_date.strftime('%Y-%m-%d')}"
    path = RAW_DATA_DIR / folder_name
    if path.exists():
        return path

    # Try next business day folder (files are dated T+1)
    from pandas.tseries.holiday import USFederalHolidayCalendar
    from pandas.tseries.offsets import CustomBusinessDay
    US_BD = CustomBusinessDay(calendar=USFederalHolidayCalendar())
    next_bd = (pd.Timestamp(report_date) + US_BD).date()
    folder_name = f"SEI {next_bd.strftime('%Y-%m-%d')}"
    path = RAW_DATA_DIR / folder_name
    if path.exists():
        return path

    # Fall back to the most recent SEI folder that exists
    all_folders = sorted(
        [p for p in RAW_DATA_DIR.glob("SEI *") if p.is_dir()],
        reverse=True,
    )
    if all_folders:
        latest = all_folders[0]
        print(f"  [INFO] Using latest available SEI folder: {latest.name}")
        return latest

    raise FileNotFoundError(f"No SEI folder found for date {report_date}")


def _find_file(folder: Path, pattern: str) -> Path:
    """Find a file matching glob pattern in folder."""
    matches = list(folder.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching '{pattern}' in {folder}")
    return max(matches, key=lambda p: p.stat().st_mtime)


# ===========================================================================
# DATA LOADERS
# ===========================================================================

def load_fund_mapping() -> pd.DataFrame:
    """Load the master fund mapping file."""
    df = pd.read_csv(MAP_FUND_PATH, encoding="latin1")
    # Keep only US-domiciled funds with a valid ticker
    df = df[df["DOMICILE"] == "US"].copy()
    df = df[df["FUND"].notna() & (df["FUND"] != "")].copy()
    df["FUND"] = df["FUND"].str.strip()
    return df


def load_attribute_mapping(report_date: date) -> dict:
    """
    Load ticker -> Attribute mapping from today's Port Review file.
    Falls back to the CSV sample if the Excel isn't available yet.
    """
    # Try the existing Port Review Excel
    output_dir = Path(r"X:\PM & Operations\Portfolio Management\Portfolio Review")
    excel_path = output_dir / f"Port_Review_{report_date.strftime('%Y%m%d')}.xlsx"

    try:
        excel_available = excel_path.exists()
    except OSError as e:
        print(f"  [WARN] Could not check attribute workbook on network drive: {e}")
        excel_available = False

    if excel_available:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
            ws = wb["Port Review Master Table"]
            mapping = {}
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                ticker, attr = row
                if ticker and attr:
                    mapping[str(ticker).strip().upper()] = str(attr).strip()
            wb.close()
            return mapping
        except Exception:
            pass

    # Fallback: use the CSV in workspace
    csv_path = Path("Port Review Master Table.csv")
    if csv_path.exists():
        df = pd.read_csv(csv_path, usecols=["Fund Ticker", "Attribute"])
        return dict(zip(df["Fund Ticker"].str.strip().str.upper(), df["Attribute"].str.strip()))

    return {}


def load_assets(sei_folder: Path) -> pd.DataFrame:
    """Load fund-level AUM from the blackbar Assets tab."""
    pattern = "*ETF_blackbar_AssetsTab*"
    path = _find_file(sei_folder, pattern)
    df = pd.read_excel(path, usecols=["ticker", "Assets"])
    df["ticker"] = df["ticker"].str.strip().str.upper()
    df = df.rename(columns={"ticker": "Fund Ticker", "Assets": "Assets"})
    return df.drop_duplicates(subset="Fund Ticker")


def load_holdings(sei_folder: Path, holdings_date: str) -> pd.DataFrame:
    """
    Load SEI tradedate holdings CSV.
    Tries the exact date first, then falls back to any holdings file in the folder.
    """
    pattern = f"SEI_GLX_Tradedate_Holdings_{holdings_date}*"
    csv_matches = list(sei_folder.glob(pattern.replace("*", ".csv")))
    if not csv_matches:
        csv_matches = list(sei_folder.glob(pattern))

    # Fall back: any holdings file in the folder (latest by mtime)
    if not csv_matches:
        all_holdings = list(sei_folder.glob("SEI_GLX_Tradedate_Holdings_*.csv"))
        if not all_holdings:
            all_holdings = list(sei_folder.glob("SEI_GLX_Tradedate_Holdings_*.xlsx"))
        if all_holdings:
            csv_matches = [max(all_holdings, key=lambda p: p.stat().st_mtime)]
            print(f"  [INFO] Using latest holdings file: {csv_matches[0].name}")

    if not csv_matches:
        raise FileNotFoundError(f"No holdings file in {sei_folder}")

    path = csv_matches[0]
    if path.suffix == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    return df


def load_cfc_holdings(sei_folder: Path) -> pd.DataFrame:
    """Load CFC Holdings file for CFC Cash calculation."""
    pattern = "Global X CFC Holdings*"
    try:
        path = _find_file(sei_folder, pattern)
    except FileNotFoundError:
        candidates = [
            p for p in sei_folder.glob("GlobalX_Positions_*.xlsx")
            if "_Option" not in p.name and "_VY" not in p.name
        ]
        if not candidates:
            return pd.DataFrame(columns=["Fund Ticker", "market_value_usd"])

        positions = pd.read_excel(max(candidates, key=lambda p: p.stat().st_mtime))
        desc = positions.get("Security Description (Short)", pd.Series("", index=positions.index))
        cfc_rows = positions[
            desc.fillna("").astype(str).str.upper().str.contains(r"\bCFC\b", regex=True)
            & positions.get("Asset Group", pd.Series("", index=positions.index)).fillna("").astype(str).str.upper().eq("MUTUAL FUND")
        ].copy()
        if cfc_rows.empty:
            return pd.DataFrame(columns=["Fund Ticker", "market_value_usd"])

        fund_map = load_fund_mapping()
        cfc_rows["Fund Ticker"] = _map_fund_tickers(
            cfc_rows,
            fund_map,
            id_col="Account Number",
            name_col="Account Name",
        )
        cfc_rows["market_value_usd"] = pd.to_numeric(
            cfc_rows["Market Value (Base)"],
            errors="coerce",
        ).fillna(0.0)
        return cfc_rows[["Fund Ticker", "market_value_usd"]]

    df = pd.read_excel(path)
    # CFC cash = sum of Cash rows per account
    cash_rows = df[df["Asset Group"].str.strip().str.upper() == "CA"]
    if cash_rows.empty:
        # Try filtering by security description
        cash_rows = df[df["Security Description (Short)"].str.upper().str.contains("CASH", na=False)]
    # Calculate USD market value: Shares/Par * Previous Price * Exchange Rate / Contract Size
    if not cash_rows.empty and "Shares/Par" in cash_rows.columns:
        cash_rows = cash_rows.copy()
        cash_rows["market_value_usd"] = (
            cash_rows["Shares/Par"].astype(float)
            * cash_rows["Previous Price"].astype(float)
            * cash_rows["Exchange Rate"].astype(float)
            / cash_rows["Contract Size"].astype(float)
        )
        return cash_rows[["Account Number", "market_value_usd"]]
    return pd.DataFrame(columns=["Account Number", "market_value_usd"])


# ===========================================================================
# INDEX DATA (Snowflake)
# ===========================================================================

def load_bbh_cash(sei_folder: Path, fund_mapping: pd.DataFrame) -> pd.DataFrame:
    """
    Load Custody Cash (USD) BBH and Foreign Ccy BBH from PROJECTED_CASH_BALANCE.
    The file is in the SEI folder: 188834_PROJECTED_CASH_BALANCE.*.xlsx

    Returns DataFrame with: Fund Ticker, Custody Cash (USD) BBH, Foreign Ccy BBH
    """
    try:
        path = _find_file(sei_folder, "188834_PROJECTED_CASH_BALANCE*")
    except FileNotFoundError:
        print("  [WARN] PROJECTED_CASH_BALANCE file not found")
        return pd.DataFrame(columns=["Fund Ticker", "Custody Cash (USD) BBH", "Foreign Ccy BBH"])

    df = pd.read_excel(path)

    # The file has: Currency Account Name, Currency Code, Value Date,
    # Opening Available + CMS Sweep Return, Projected Closing Available Balance
    # We need T+0 (today's value date = the earliest date in the file)
    # Parse Value Date
    df["Value Date"] = pd.to_datetime(df["Value Date"], format="%m/%d/%Y", errors="coerce")
    t0_date = df["Value Date"].min()
    today_rows = df[df["Value Date"] == t0_date].copy()

    # Extract fund ticker from Currency Account Name
    # Pattern: "DKK NORW" -> NORW, "NORW" -> NORW, "NORW OLD" -> NORW OLD
    # The USD row typically has just the fund name (e.g., "NORW")
    # Non-USD rows have "CCY FUND" format (e.g., "NOK NORW")
    # Build BBH account -> ticker mapping from fund_mapping
    bbh_accts = fund_mapping[["BBH_ACCOUNT_NUM", "FUND"]].dropna(subset=["BBH_ACCOUNT_NUM"])
    bbh_accts = bbh_accts[bbh_accts["BBH_ACCOUNT_NUM"].apply(
        lambda x: str(x).replace(".", "").isdigit()
    )]
    bbh_acct_to_ticker = dict(zip(
        bbh_accts["BBH_ACCOUNT_NUM"].astype(float).astype(int),
        bbh_accts["FUND"]
    ))

    # Map Head Account Number to ticker
    today_rows = today_rows.copy()
    today_rows["Head Account Number"] = pd.to_numeric(today_rows["Head Account Number"], errors="coerce")
    today_rows["Fund Ticker"] = today_rows["Head Account Number"].map(bbh_acct_to_ticker)

    # Parse the cash amounts (they may be strings with commas)
    def _parse_amount(val):
        if pd.isna(val):
            return 0.0
        s = str(val).replace(",", "").strip()
        if s in ("", ".00", "0"):
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0

    today_rows["amount"] = today_rows["Opening Available + CMS Sweep Return"].apply(_parse_amount)

    # USD rows -> Custody Cash (USD) BBH (sum per fund)
    usd_rows = today_rows[today_rows["Currency Code"] == "USD"]
    bbh_usd = usd_rows.groupby("Fund Ticker")["amount"].sum().reset_index()
    bbh_usd.columns = ["Fund Ticker", "Custody Cash (USD) BBH"]

    # Non-USD rows -> Foreign Ccy BBH (sum per fund, already in reporting currency = USD)
    non_usd = today_rows[today_rows["Currency Code"] != "USD"]
    bbh_fcy = non_usd.groupby("Fund Ticker")["amount"].sum().reset_index()
    bbh_fcy.columns = ["Fund Ticker", "Foreign Ccy BBH"]

    # Merge
    result = bbh_usd.merge(bbh_fcy, on="Fund Ticker", how="outer").fillna(0)
    return result


def load_bny_cash(report_date: date, fund_tickers: list) -> pd.DataFrame:
    """
    Load Custody Cash (USD) BNY and Foreign Ccy BNY from per-fund BNY Cash Balance files.

    Formula: USD Ending Balance (T+0) + Sweep Vehicle Ending Balance
    Foreign Ccy: sum of non-USD Ending Balance Reporting Currency for T+0

    Returns DataFrame with: Fund Ticker, Custody Cash (USD) BNY, Foreign Ccy BNY
    """
    from pandas.tseries.holiday import USFederalHolidayCalendar
    from pandas.tseries.offsets import CustomBusinessDay
    US_BD = CustomBusinessDay(calendar=USFederalHolidayCalendar())

    # BNY Cash Balance folder is dated T+1
    folder_date = (pd.Timestamp(report_date) + US_BD).date()
    folder = BNY_CASH_DIR / folder_date.strftime("%Y%m%d")

    if not folder.exists():
        # Fall back to the latest dated folder
        all_folders = sorted(
            [p for p in BNY_CASH_DIR.glob("[0-9]*") if p.is_dir()],
            reverse=True,
        )
        if all_folders:
            folder = all_folders[0]
            print(f"  [INFO] Using latest BNY Cash Balance folder: {folder.name}")
        else:
            print(f"  [WARN] No BNY Cash Balance folders found")
            return pd.DataFrame(columns=["Fund Ticker", "Custody Cash (USD) BNY", "Foreign Ccy BNY"])

    results = []
    for ticker in fund_tickers:
        # Find the file for this ticker
        pattern = f"{ticker} BNY Cash Balance*"
        matches = list(folder.glob(pattern))
        if not matches:
            results.append({"Fund Ticker": ticker, "Custody Cash (USD) BNY": 0.0, "Foreign Ccy BNY": 0.0})
            continue

        try:
            df = pd.read_excel(matches[0])

            # Find T+0 date (the earliest Cash Reporting Date)
            df["Cash Reporting Date"] = pd.to_datetime(df["Cash Reporting Date"], errors="coerce")
            t0_date = df["Cash Reporting Date"].min()
            today = df[df["Cash Reporting Date"] == t0_date].copy()

            # USD: Ending Balance + Sweep
            usd_rows = today[today["Local Currency Code"] == "USD"]
            usd_ending = usd_rows["Ending Balance Reporting Currency"].sum()

            # Sweep vehicle (has NaT for Cash Reporting Date but has Sweep Vehicle Number)
            sweep_rows = df[df["Sweep Vehicle Number"].notna() & (df["Local Currency Code"] == "USD")]
            sweep_amount = sweep_rows["Ending Balance Reporting Currency"].sum()

            bny_usd = usd_ending + sweep_amount

            # Foreign Ccy: sum of non-USD ending balances for T+0
            non_usd = today[today["Local Currency Code"] != "USD"]
            fcy_bny = non_usd["Ending Balance Reporting Currency"].sum()

            results.append({
                "Fund Ticker": ticker,
                "Custody Cash (USD) BNY": bny_usd,
                "Foreign Ccy BNY": fcy_bny,
            })
        except Exception as e:
            results.append({"Fund Ticker": ticker, "Custody Cash (USD) BNY": 0.0, "Foreign Ccy BNY": 0.0})

    return pd.DataFrame(results)


def load_actual_cash(sei_folder: Path, fund_mapping: pd.DataFrame) -> pd.DataFrame:
    """
    Load Actual Cash from GlobalX_Positions file.
    Actual Cash = sum of Cash + Currency rows' Market Value (Base) per fund.

    Returns DataFrame with: Fund Ticker, Actual Cash $
    """
    # Use the main positions file (not the _Option_VY variant)
    candidates = [
        p for p in sei_folder.glob("GlobalX_Positions_*.xlsx")
        if "_Option" not in p.name and "_VY" not in p.name
    ]
    if not candidates:
        print("  [WARN] GlobalX_Positions file not found")
        return pd.DataFrame(columns=["Fund Ticker", "Actual Cash $"])
    path = max(candidates, key=lambda p: p.stat().st_mtime)

    df = pd.read_excel(path)

    df["Fund Ticker"] = _map_fund_tickers(
        df,
        fund_mapping,
        id_col="Account Number",
        name_col="Account Name",
    )
    missing_ticker = df["Fund Ticker"].isna()
    if missing_ticker.any():
        valid = fund_mapping[fund_mapping["FUND_NUMBER"].notna()].copy()
        id_to_ticker = dict(zip(valid["FUND_NUMBER"].astype(int).astype(str), valid["FUND"]))
        acct_prefix = (
            pd.to_numeric(df.loc[missing_ticker, "Account Number"], errors="coerce")
            .dropna()
            .astype(int)
            .astype(str)
            .str[:4]
        )
        df.loc[acct_prefix.index, "Fund Ticker"] = acct_prefix.map(id_to_ticker)
    df = df[df["Fund Ticker"].notna()].copy()
    if "Account Name" in df.columns:
        df = df[~df["Account Name"].fillna("").astype(str).str.upper().str.contains("CFC")].copy()

    # Cash + Currency rows (both contribute to Actual Cash)
    df["Asset Group Clean"] = df["Asset Group"].str.strip()
    cash_rows = df[df["Asset Group Clean"].isin(["Cash", "Currency"])].copy()

    # Sum Market Value (Base) per fund
    cash_by_fund = cash_rows.groupby("Fund Ticker")["Market Value (Base)"].sum().reset_index()
    cash_by_fund.columns = ["Fund Ticker", "Actual Cash $"]

    return cash_by_fund


def load_futures(holdings: pd.DataFrame, fund_mapping: pd.DataFrame) -> pd.DataFrame:
    """
    Load futures market value from SEI tradedate holdings.
    Futures rows have security_type = 'Future (New)'.

    Returns DataFrame with: Fund Ticker, Futures $, Futures PNA
    """
    futures = holdings[holdings["security_type"] == "Future (New)"].copy()
    if futures.empty:
        return pd.DataFrame(columns=["Fund Ticker", "Futures $", "Futures PNA"])

    futures["Fund Ticker"] = _map_fund_tickers(futures, fund_mapping)
    futures = futures[futures["Fund Ticker"].notna()].copy()

    futures["market_value_num"] = pd.to_numeric(futures["market_value"], errors="coerce").fillna(0.0)
    futures["pna"] = pd.to_numeric(futures["percent_of_net_assets"], errors="coerce").fillna(0.0)
    fut_by_fund = futures.groupby("Fund Ticker").agg({
        "market_value_num": "sum",
        "pna": "sum",
    }).reset_index()
    fut_by_fund.columns = ["Fund Ticker", "Futures $", "Futures PNA"]
    fut_by_fund["Futures PNA"] = fut_by_fund["Futures PNA"] / 100.0

    return fut_by_fund


def load_pending_bny(report_date: date, fund_tickers: list) -> pd.DataFrame:
    """
    Load Pending (USD) BNY from BNY_Unsettled_Trades file via SFTP.
    Source: sftp.globalxetfs.com /bnymellon/inbound/BNY_Unsettled_Trades_DDMMYYYYHHMMSS.csv

    Uses the latest file for the report date.
    Pending Amount = sum of Local Net Amount for USD trades per fund / Assets.

    Returns DataFrame with: Fund Ticker, Pending (USD) BNY (as fraction of Assets — 
    NOTE: currently returns raw dollar sum; division by Assets happens in caller)
    """
    import paramiko

    SFTP_HOST = "sftp.globalxetfs.com"
    SFTP_PORT = 22
    SFTP_USER = "gxpapm"
    SFTP_PASS = "1ge3isTfwswhBE8jJLgC"
    REMOTE_DIR = "/bnymellon/inbound"

    # Date format in filename: DDMMYYYY
    date_prefix = report_date.strftime("%d%m%Y")

    try:
        transport = paramiko.Transport((SFTP_HOST, SFTP_PORT))
        transport.connect(username=SFTP_USER, password=SFTP_PASS)
        sftp = paramiko.SFTPClient.from_transport(transport)

        # Find the latest file for this date
        all_files = sftp.listdir(REMOTE_DIR)
        day_files = sorted([f for f in all_files if f"Unsettled_Trades_{date_prefix}" in f])

        if not day_files:
            print(f"  [WARN] No unsettled trades file found for {date_prefix} on SFTP")
            sftp.close()
            transport.close()
            return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending (USD) BNY": 0.0})

        # Use the latest (last) file of the day
        target_file = day_files[-1]
        local_path = Path(f"BNY_Unsettled_Trades_{date_prefix}.csv")
        sftp.get(f"{REMOTE_DIR}/{target_file}", str(local_path))
        print(f"  [OK] Downloaded unsettled trades: {target_file}")

        sftp.close()
        transport.close()

    except Exception as e:
        print(f"  [WARN] SFTP download failed: {e}")
        return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending (USD) BNY": 0.0})

    # Parse the file
    try:
        df = pd.read_csv(local_path)

        def _parse_amount(val):
            s = str(val).strip().replace(",", "")
            if s.startswith("(") and s.endswith(")"):
                return -float(s[1:-1])
            return float(s)

        df["amount_usd"] = df["Local Net Amount"].apply(_parse_amount)

        # Map Account Number to ticker using fund mapping
        fund_map = load_fund_mapping()
        bny_accts = fund_map[["BNY_ACCOUNT", "FUND"]].dropna(subset=["BNY_ACCOUNT"])
        bny_accts = bny_accts[bny_accts["BNY_ACCOUNT"].apply(
            lambda x: str(x).replace(".", "").isdigit()
        )]
        acct_to_ticker = dict(zip(
            bny_accts["BNY_ACCOUNT"].astype(float).astype(int),
            bny_accts["FUND"]
        ))

        df["Fund Ticker"] = df["Account Number"].map(acct_to_ticker)
        mapped = df[df["Fund Ticker"].notna() & (df["Local Currency Code"] == "USD")]

        # Sum pending amount by fund
        # TODO: Confirm filtering logic (all rows? only failed? only past settlement?)
        pending = mapped.groupby("Fund Ticker")["amount_usd"].sum().reset_index()
        pending.columns = ["Fund Ticker", "Pending (USD) BNY"]

        # Fill missing tickers with 0
        all_tickers = pd.DataFrame({"Fund Ticker": fund_tickers})
        result = all_tickers.merge(pending, on="Fund Ticker", how="left")
        result["Pending (USD) BNY"] = result["Pending (USD) BNY"].fillna(0)

        return result

    except Exception as e:
        print(f"  [WARN] Failed to parse unsettled trades: {e}")
        return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending (USD) BNY": 0.0})


def load_pending_cil(report_date: date, fund_tickers: list) -> pd.DataFrame:
    """
    Load Pending CIL from the SEI Final Blotter email attachment.
    The file is SEI_GlobalX_MMDDYYYY.csv attached to the 'Global X Final Blotter' email.

    Looks for the file in the SEI folder first, then tries Outlook.
    Returns DataFrame with: Fund Ticker, Pending CIL (raw dollar amount)
    """
    # Try to find the blotter file locally in the SEI folder
    from pandas.tseries.holiday import USFederalHolidayCalendar
    from pandas.tseries.offsets import CustomBusinessDay
    US_BD = CustomBusinessDay(calendar=USFederalHolidayCalendar())
    sei_folder_date = (pd.Timestamp(report_date) + US_BD).date()
    sei_folder = RAW_DATA_DIR / f"SEI {sei_folder_date.strftime('%Y-%m-%d')}"

    date_str = report_date.strftime("%m%d%Y")
    blotter_pattern = f"SEI_GlobalX_{date_str}*"

    local_path = None
    if sei_folder.exists():
        matches = list(sei_folder.glob(blotter_pattern))
        if matches:
            local_path = matches[0]

    # Also check a common download location
    if local_path is None:
        downloads = Path.home() / "Downloads"
        matches = list(downloads.glob(blotter_pattern))
        if matches:
            local_path = max(matches, key=lambda p: p.stat().st_mtime)

    if local_path is None:
        # Try to pull from Outlook
        try:
            local_path = _extract_blotter_from_outlook(report_date)
        except Exception as e:
            print(f"  [WARN] Could not get SEI blotter: {e}")

    if local_path is None or not local_path.exists():
        print(f"  [WARN] SEI blotter file not found for {date_str}")
        return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending CIL": 0.0})

    try:
        df = pd.read_csv(local_path)
        print(f"  [OK] Loaded SEI blotter: {local_path.name} ({len(df)} rows)")

        # The blotter should have a ticker column and a CIL/pending amount
        # Look for columns that match
        cols_lower = {c.lower(): c for c in df.columns}

        # Find ticker column
        ticker_col = None
        for candidate in ["ticker", "fund", "fund_ticker", "etf"]:
            if candidate in cols_lower:
                ticker_col = cols_lower[candidate]
                break

        # Find CIL amount column
        cil_col = None
        for candidate in ["pending cil", "cil", "cil_amount", "pending_cil", "net_amount"]:
            if candidate in cols_lower:
                cil_col = cols_lower[candidate]
                break

        if ticker_col is None or cil_col is None:
            print(f"  [WARN] Could not identify columns in blotter. Columns: {list(df.columns)}")
            return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending CIL": 0.0})

        # Sum CIL by ticker
        df[cil_col] = pd.to_numeric(df[cil_col], errors="coerce").fillna(0)
        cil_by_fund = df.groupby(ticker_col)[cil_col].sum().reset_index()
        cil_by_fund.columns = ["Fund Ticker", "Pending CIL"]
        cil_by_fund["Fund Ticker"] = cil_by_fund["Fund Ticker"].str.strip().str.upper()

        all_tickers = pd.DataFrame({"Fund Ticker": fund_tickers})
        result = all_tickers.merge(cil_by_fund, on="Fund Ticker", how="left")
        result["Pending CIL"] = result["Pending CIL"].fillna(0)
        return result

    except Exception as e:
        print(f"  [WARN] Failed to parse SEI blotter: {e}")
        return pd.DataFrame({"Fund Ticker": fund_tickers, "Pending CIL": 0.0})


def _extract_blotter_from_outlook(report_date: date) -> Optional[Path]:
    """Try to extract the SEI blotter attachment from today's Outlook inbox."""
    try:
        import win32com.client
    except ImportError:
        return None

    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)  # 6 = Inbox

    date_str = report_date.strftime("%m%d%Y")
    target_name = f"SEI_GlobalX_{date_str}"

    # Search recent messages
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)

    for i in range(min(50, messages.Count)):
        msg = messages.Item(i + 1)
        if "Final Blotter" in str(msg.Subject):
            for att in msg.Attachments:
                if target_name in att.FileName:
                    save_path = Path(f"./{att.FileName}")
                    att.SaveAsFile(str(save_path.resolve()))
                    print(f"  [OK] Extracted blotter from Outlook: {att.FileName}")
                    return save_path

    return None


def _clean_code_list(values) -> list:
    """Return SQL-safe-ish index codes from mapFUND fields."""
    cleaned = []
    for val in values:
        if pd.isna(val):
            continue
        s = str(val).strip()
        if not s or s in {"--", "0", "nan", "None"}:
            continue
        cleaned.append(s.replace("'", "''"))
    return sorted(set(cleaned))


def _normalize_index_weights(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize provider constituent rows to decimal weights."""
    if df.empty:
        return df
    out = df.copy()
    out["INDEX_WEIGHTING"] = pd.to_numeric(
        out["INDEX_WEIGHTING"].astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False),
        errors="coerce",
    )
    # Provider feeds are mixed: some are decimals, some are percentage points.
    # Normalize per index because percent-weight feeds can have each holding below 1
    # while still summing near 100.
    for _, idx in out.groupby("INDEX").groups.items():
        weights = out.loc[idx, "INDEX_WEIGHTING"].dropna().abs()
        if weights.empty:
            continue
        if weights.max() > 1.0 or weights.sum() > 2.0:
            out.loc[idx, "INDEX_WEIGHTING"] = out.loc[idx, "INDEX_WEIGHTING"] / 100.0
    for col in ["SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER"]:
        if col not in out.columns:
            out[col] = np.nan
    return out[["INDEX", "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER", "INDEX_WEIGHTING", "DATE"]]


def load_index_constituents(fund_mapping: pd.DataFrame, as_of_date: date) -> pd.DataFrame:
    """
    Load benchmark constituents from non-legacy provider feeds.

    Returns normalized rows with INDEX, SECURITY_SEDOL, SECURITY_ISIN,
    SECURITY_TICKER, INDEX_WEIGHTING (decimal), and DATE.
    """
    frames = []
    conn = _sf_connect()
    try:
        cur = conn.cursor()

        # Solactive composite also carries many INDXX-style index codes used by US ETFs.
        # Use closing constituents because SEI tradedate holdings are end-of-day.
        sol_codes = _clean_code_list(
            pd.concat([
                fund_mapping.get("IDX_CONSTITUENTS", pd.Series(dtype=object)),
                fund_mapping.get("IDX_CODE", pd.Series(dtype=object)),
                fund_mapping.get("IDX_VALUES", pd.Series(dtype=object)),
            ])
        )
        if sol_codes:
            codes_str = ",".join(f"'{c}'" for c in sol_codes)
            cur.execute(f"""
                WITH latest AS (
                    SELECT "INDEX", MAX(DATA_DATE) AS latest_date
                    FROM PROD.SOLACTIVE.CLOSING_COMPOSITE
                    WHERE "INDEX" IN ({codes_str})
                      AND TRY_TO_DATE(DATA_DATE) <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                    GROUP BY "INDEX"
                )
                SELECT s."INDEX",
                       s.SECURITY_SEDOL,
                       s.ISIN AS SECURITY_ISIN,
                       s.SECURITY_TICKER,
                       s.INDEX_WEIGHTING,
                       TRY_TO_DATE(s.DATA_DATE) AS DATE
                FROM PROD.SOLACTIVE.CLOSING_COMPOSITE s
                INNER JOIN latest l
                  ON s."INDEX" = l."INDEX"
                 AND TRY_TO_DATE(s.DATA_DATE) = l.latest_date
                WHERE s.INDEX_WEIGHTING IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # S&P feeds are split by fund/index family into *_CLOSING tables.
        cur.execute("""
            SELECT table_name
            FROM PROD.INFORMATION_SCHEMA.TABLES
            WHERE table_schema = 'S_AND_P'
              AND table_name LIKE '%CLOSING'
        """)
        sp_tables = {r[0].upper() for r in cur.fetchall() if r[0].upper().endswith("_CLOSING")}
        sp_funds = fund_mapping[fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("S&P")].copy()
        for _, row in sp_funds.iterrows():
            fund = str(row["FUND"]).strip().upper()
            table = f"{fund}_CLOSING"
            if table not in sp_tables:
                continue
            idx_code = row.get("IDX_CONSTITUENTS")
            if pd.isna(idx_code) or str(idx_code).strip() in {"", "--", "0"}:
                idx_code = row.get("IDX_CODE")
            idx_code = str(idx_code).strip()
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(EFFECTIVE_DATE) AS latest_date
                    FROM PROD.S_AND_P.{table}
                    WHERE EFFECTIVE_DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                )
                SELECT '{idx_code.replace("'", "''")}' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       ISIN AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       INDEX_WEIGHT AS INDEX_WEIGHTING,
                       EFFECTIVE_DATE AS DATE
                FROM PROD.S_AND_P.{table}
                WHERE EFFECTIVE_DATE = (SELECT latest_date FROM latest)
                  AND INDEX_WEIGHT IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        sp500_aliases = set()
        for col in ["IDX_CONSTITUENTS", "IDX_CODE"]:
            if col in fund_mapping.columns:
                values = fund_mapping.loc[
                    fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("S&P"),
                    col,
                ].dropna().astype(str).str.strip()
                sp500_aliases.update(
                    v for v in values
                    if v.upper() in {"S&P 500", "SP500", "SPX", "US500"} or v == "500"
                )
        if sp500_aliases:
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(EFFECTIVE_DATE) AS latest_date
                    FROM PROD.S_AND_P.SP500_CLOSING
                    WHERE EFFECTIVE_DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                )
                SELECT '__SP500__' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       ISIN AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       INDEX_WEIGHT AS INDEX_WEIGHTING,
                       EFFECTIVE_DATE AS DATE
                FROM PROD.S_AND_P.SP500_CLOSING
                WHERE EFFECTIVE_DATE = (SELECT latest_date FROM latest)
                  AND INDEX_WEIGHT IS NOT NULL
            """)
            sp500 = cur.fetch_pandas_all()
            for alias in sorted(sp500_aliases):
                alias_rows = sp500.copy()
                alias_rows["INDEX"] = alias
                frames.append(alias_rows)

        # DAX closing constituents.
        dax_codes = _clean_code_list(fund_mapping.loc[
            fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("STOXX"),
            "IDX_CONSTITUENTS"
        ])
        if dax_codes:
            codes_str = ",".join(f"'{c}'" for c in dax_codes)
            cur.execute(f"""
                WITH latest AS (
                    SELECT INDEX_SYMBOL, MAX(DATE) AS latest_date
                    FROM PROD.DAX.DAX_INDEX_CLOSE
                    WHERE INDEX_SYMBOL IN ({codes_str})
                      AND DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                    GROUP BY INDEX_SYMBOL
                )
                SELECT d.INDEX_SYMBOL AS "INDEX",
                       NULL AS SECURITY_SEDOL,
                       d.ISIN AS SECURITY_ISIN,
                       NULL AS SECURITY_TICKER,
                       d.WEIGHT AS INDEX_WEIGHTING,
                       d.DATE AS DATE
                FROM PROD.DAX.DAX_INDEX_CLOSE d
                INNER JOIN latest l
                  ON d.INDEX_SYMBOL = l.INDEX_SYMBOL
                 AND d.DATE = l.latest_date
                WHERE d.WEIGHT IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # MSCI EOD constituents for numeric mapFUND benchmark codes.
        msci_values = []
        if "IDX_PROVIDER" in fund_mapping.columns and "IDX_CONSTITUENTS" in fund_mapping.columns:
            msci_values = fund_mapping.loc[
                fund_mapping["IDX_PROVIDER"].astype(str).str.upper().eq("MSCI"),
                "IDX_CONSTITUENTS",
            ].dropna().astype(str).str.strip().tolist()
        msci_codes = sorted({str(int(float(v))) for v in msci_values if v.replace(".", "", 1).isdigit()})
        if msci_codes:
            codes_str = ",".join(msci_codes)
            cur.execute(f"""
                WITH latest AS (
                    SELECT MSCI_INDEX_CODE, MAX(CALCULATION_DATE) AS latest_date
                    FROM PROD.MSCI.INDEX_EOD
                    WHERE MSCI_INDEX_CODE IN ({codes_str})
                      AND CALCULATION_DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                    GROUP BY MSCI_INDEX_CODE
                )
                SELECT TO_VARCHAR(m.MSCI_INDEX_CODE) AS "INDEX",
                       m.SECURITY_SEDOL,
                       m.ISIN AS SECURITY_ISIN,
                       m.SECURITY_TICKER,
                       m.INDEX_WEIGHTING,
                       m.CALCULATION_DATE AS DATE
                FROM PROD.MSCI.INDEX_EOD m
                INNER JOIN latest l
                  ON m.MSCI_INDEX_CODE = l.MSCI_INDEX_CODE
                 AND m.CALCULATION_DATE = l.latest_date
                WHERE m.INDEX_WEIGHTING IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # Mirae custom US close constituents.
        mirae_codes = _clean_code_list(fund_mapping.loc[
            fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("MIRAE"),
            "IDX_CONSTITUENTS"
        ])
        if mirae_codes:
            codes_str = ",".join(f"'{c}'" for c in mirae_codes)
            cur.execute(f"""
                WITH latest AS (
                    SELECT INDEX_TICKER, MAX(DATA_DATE::DATE) AS latest_date
                    FROM PROD.MIRAE.CONSOLIDATED_US_CLOSE
                    WHERE INDEX_TICKER IN ({codes_str})
                      AND DATA_DATE::DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                    GROUP BY INDEX_TICKER
                )
                SELECT m.INDEX_TICKER AS "INDEX",
                       m.SECURITY_SEDOL,
                       m.ISIN AS SECURITY_ISIN,
                       m.SECURITY_TICKER,
                       m.INDEX_WEIGHTING,
                       m.DATA_DATE::DATE AS DATE
                FROM PROD.MIRAE.CONSOLIDATED_US_CLOSE m
                INNER JOIN latest l
                  ON m.INDEX_TICKER = l.INDEX_TICKER
                 AND m.DATA_DATE::DATE = l.latest_date
                WHERE m.INDEX_WEIGHTING IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # NASDAQ closing constituents, used for NDX-based option products.
        nasdaq_codes = _clean_code_list(fund_mapping.loc[
            fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("NASDAQ"),
            "IDX_CONSTITUENTS"
        ])
        if nasdaq_codes:
            codes_str = ",".join(f"'{c}'" for c in nasdaq_codes)
            cur.execute(f"""
                WITH latest AS (
                    SELECT INDEX_SYMBOL, MAX(DATA_DATE::DATE) AS latest_date
                    FROM PROD.NASDAQ.NASDAQ_CLOSING
                    WHERE INDEX_SYMBOL IN ({codes_str})
                      AND DATA_DATE::DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                    GROUP BY INDEX_SYMBOL
                )
                SELECT n.INDEX_SYMBOL AS "INDEX",
                       n.SEDOL AS SECURITY_SEDOL,
                       n.ISIN AS SECURITY_ISIN,
                       n.SYMBOL AS SECURITY_TICKER,
                       n.INDEX_WEIGHT AS INDEX_WEIGHTING,
                       n.DATA_DATE::DATE AS DATE
                FROM PROD.NASDAQ.NASDAQ_CLOSING n
                INNER JOIN latest l
                  ON n.INDEX_SYMBOL = l.INDEX_SYMBOL
                 AND n.DATA_DATE::DATE = l.latest_date
                WHERE n.INDEX_WEIGHT IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        baml_codes = set()
        if "IDX_CONSTITUENTS" in fund_mapping.columns:
            baml_codes.update(
                fund_mapping.loc[
                    fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().isin({"ICE", "BAML"}),
                    "IDX_CONSTITUENTS",
                ].dropna().astype(str).str.strip()
            )
        if "GDM" in baml_codes:
            cur.execute(f"""
                WITH latest AS (
                SELECT MAX(AS_OF_DATE) AS latest_date
                    FROM PROD.BAML.GDM_EOD
                    WHERE AS_OF_DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                      AND INDEX_TICKER = 'GDM'
                )
                SELECT 'GDM' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       ISIN_NUMBER AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       WEIGHT AS INDEX_WEIGHTING,
                       AS_OF_DATE AS DATE
                FROM PROD.BAML.GDM_EOD
                WHERE AS_OF_DATE = (SELECT latest_date FROM latest)
                  AND INDEX_TICKER = 'GDM'
                  AND WEIGHT IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        if "NYSE100T" in baml_codes or "NYSE100" in baml_codes:
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(AS_OF_DATE) AS latest_date
                    FROM PROD.BAML.NYSE100_CLOSING
                    WHERE AS_OF_DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                      AND INDEX_TICKER = 'NYSE100'
                )
                SELECT 'NYSE100T' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       ISIN_NUMBER AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       WEIGHT AS INDEX_WEIGHTING,
                       AS_OF_DATE AS DATE
                FROM PROD.BAML.NYSE100_CLOSING
                WHERE AS_OF_DATE = (SELECT latest_date FROM latest)
                  AND INDEX_TICKER = 'NYSE100'
                  AND WEIGHT IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        if "PFTF" in baml_codes:
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(FILE_MODIFICATION_TIME) AS latest_time
                    FROM PROD.BAML.PARALLEL_PFTF_DAILY
                    WHERE FILE_MODIFICATION_TIME::DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                      AND INDEX_NAME = 'PFTF'
                )
                SELECT 'PFTF' AS "INDEX",
                       NULL AS SECURITY_SEDOL,
                       ISIN_NUMBER AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       PERCENT_CURRENT_MKT_VALUE AS INDEX_WEIGHTING,
                       FILE_MODIFICATION_TIME::DATE AS DATE
                FROM PROD.BAML.PARALLEL_PFTF_DAILY
                WHERE FILE_MODIFICATION_TIME = (SELECT latest_time FROM latest)
                  AND INDEX_NAME = 'PFTF'
                  AND PERCENT_CURRENT_MKT_VALUE IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        if "PLCR4PM" in baml_codes:
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(FILE_MODIFICATION_TIME) AS latest_time
                    FROM PROD.BAML.PARALLEL_PLCR_DAILY
                    WHERE FILE_MODIFICATION_TIME::DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                      AND INDEX_NAME = 'PLCR4PM'
                )
                SELECT 'PLCR4PM' AS "INDEX",
                       NULL AS SECURITY_SEDOL,
                       ISIN_NUMBER AS SECURITY_ISIN,
                       TICKER AS SECURITY_TICKER,
                       PERCENT_CURRENT_MKT_VALUE AS INDEX_WEIGHTING,
                       FILE_MODIFICATION_TIME::DATE AS DATE
                FROM PROD.BAML.PARALLEL_PLCR_DAILY
                WHERE FILE_MODIFICATION_TIME = (SELECT latest_time FROM latest)
                  AND INDEX_NAME = 'PLCR4PM'
                  AND PERCENT_CURRENT_MKT_VALUE IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # Russell 2000 capped/open service.
        russell_codes = _clean_code_list(fund_mapping.loc[
            fund_mapping.get("IDX_PROVIDER", "").astype(str).str.upper().eq("RUSSELL"),
            "IDX_CONSTITUENTS"
        ])
        if russell_codes:
            cur.execute(f"""
                WITH latest AS (
                    SELECT MAX(DATE) AS latest_date
                    FROM PROD.FTSE_RUSSELL.RUSSELL_2000_INDEX_OPEN_CONSTITUENT_SERVICE
                    WHERE DATE <= '{as_of_date.strftime('%Y-%m-%d')}'::DATE
                )
                SELECT 'RU2000RC' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       ISIN AS SECURITY_ISIN,
                       BBG_TICKER AS SECURITY_TICKER,
                       TRY_TO_DOUBLE(REPLACE("%_Weight_in_Index", '%', '')) AS INDEX_WEIGHTING,
                       DATE
                FROM PROD.FTSE_RUSSELL.RUSSELL_2000_INDEX_OPEN_CONSTITUENT_SERVICE
                WHERE DATE = (SELECT latest_date FROM latest)
                  AND "%_Weight_in_Index" IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        # FTSE ASEAN 40 has a simple current constituent table.
        if "ASEAN40" in set(fund_mapping.get("IDX_CONSTITUENTS", pd.Series(dtype=object)).dropna().astype(str)):
            cur.execute("""
                WITH latest AS (
                    SELECT MAX(FILE_MODIFICATION_TIME) AS latest_time
                    FROM PROD.FTSE_RUSSELL.ASEAN_40_CONSTITUENTS
                    WHERE INDEX_MARKER = 'ASEAN40'
                )
                SELECT 'ASEAN40' AS "INDEX",
                       SEDOL AS SECURITY_SEDOL,
                       NULL AS SECURITY_ISIN,
                       NULL AS SECURITY_TICKER,
                       WT_FTSE_ASEAN_40_INDEX AS INDEX_WEIGHTING,
                       FILE_MODIFICATION_TIME::DATE AS DATE
                FROM PROD.FTSE_RUSSELL.ASEAN_40_CONSTITUENTS
                WHERE FILE_MODIFICATION_TIME = (SELECT latest_time FROM latest)
                  AND INDEX_MARKER = 'ASEAN40'
                  AND WT_FTSE_ASEAN_40_INDEX IS NOT NULL
            """)
            frames.append(cur.fetch_pandas_all())

        cur.close()
    finally:
        conn.close()

    if not frames:
        return pd.DataFrame(columns=[
            "INDEX", "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER",
            "INDEX_WEIGHTING", "DATE",
        ])

    normalized = [_normalize_index_weights(f) for f in frames if f is not None and not f.empty]
    if not normalized:
        return pd.DataFrame(columns=[
            "INDEX", "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER",
            "INDEX_WEIGHTING", "DATE",
        ])
    result = pd.concat(normalized, ignore_index=True)
    result = result[result["INDEX_WEIGHTING"].notna()].copy()
    result["DATE"] = pd.to_datetime(result["DATE"], errors="coerce")
    return result.drop_duplicates(subset=["INDEX", "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER"])


# ===========================================================================
# CALCULATIONS
# ===========================================================================

def _clean_join_key(series: pd.Series) -> pd.Series:
    return (
        series.fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .replace({"NAN": "", "NONE": "", "NULL": ""})
    )


def _clean_ticker_join_key(series: pd.Series) -> pd.Series:
    """Normalize Bloomberg-style tickers for fallback rank joins."""
    cleaned = (
        series.fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
        .str.replace(r"\s+(EQUITY|INDEX|GOVT|PFD|COMDTY)$", "", regex=True)
        .str.replace(r"-R(\s|$)", r"\1", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return cleaned.replace({"NAN": "", "NONE": "", "NULL": "", "0X2A": ""})


def _fund_name_lookup(fund_mapping: pd.DataFrame) -> dict:
    """Build normalized fund-name aliases used by SEI holdings/positions files."""
    lookup = {}
    for name_col in ["FUND_NAME_HOLDINGS", "FUND_NAME", "FUND_NAME_TAXLOTS"]:
        if name_col not in fund_mapping.columns:
            continue
        names = fund_mapping[fund_mapping[name_col].notna()].copy()
        lookup.update(dict(zip(
            names[name_col].astype(str).str.strip().str.upper(),
            names["FUND"].astype(str).str.strip(),
        )))
    return lookup


def _map_fund_tickers(
    df: pd.DataFrame,
    fund_mapping: pd.DataFrame,
    id_col: str = "portfolio_id",
    name_col: str = "sei_fund_nm",
) -> pd.Series:
    """Map a source file's fund id/name fields to the public fund ticker."""
    valid = fund_mapping[fund_mapping["FUND_NUMBER"].notna()].copy()
    id_to_ticker = dict(zip(valid["FUND_NUMBER"].astype(int), valid["FUND"]))

    mapped = pd.Series(np.nan, index=df.index, dtype=object)
    if id_col in df.columns:
        ids = pd.to_numeric(df[id_col], errors="coerce")
        mapped = ids.map(id_to_ticker)

    if name_col in df.columns:
        missing = mapped.isna()
        name_to_ticker = _fund_name_lookup(fund_mapping)
        mapped.loc[missing] = (
            df.loc[missing, name_col]
            .astype(str)
            .str.strip()
            .str.upper()
            .map(name_to_ticker)
        )
    return mapped


def _build_weight_comparison(
    fund_holdings: pd.DataFrame,
    idx_weights: pd.DataFrame,
    allow_ticker_fallback: bool = False,
) -> pd.DataFrame:
    """
    Side-by-side holdings/index comparison for one fund.

    actual_weight and INDEX_WEIGHTING are decimals. diff = actual - index.
    The join uses whichever identifier, SEDOL or ISIN, gives the best matched
    weight coverage for the fund/index pair.
    """
    if fund_holdings.empty or idx_weights.empty:
        return pd.DataFrame()

    actual_cols = [
        "security_sedol", "security_isin", "security_ticker",
        "security_type", "percent_of_market_value", "percent_of_net_assets",
    ]
    if "actual_weight" in fund_holdings.columns:
        actual_cols.append("actual_weight")
    actual = fund_holdings[actual_cols].copy()
    if "actual_weight" in actual.columns:
        actual["actual_weight"] = pd.to_numeric(actual["actual_weight"], errors="coerce").fillna(0.0)
    else:
        actual_pmv = pd.to_numeric(actual["percent_of_market_value"], errors="coerce")
        actual_pna = pd.to_numeric(actual["percent_of_net_assets"], errors="coerce")
        actual["actual_weight"] = actual_pna.fillna(actual_pmv).fillna(0.0) / 100.0

    index = idx_weights[[
        "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER", "INDEX_WEIGHTING"
    ]].copy()
    index["INDEX_WEIGHTING"] = pd.to_numeric(index["INDEX_WEIGHTING"], errors="coerce").fillna(0.0)

    candidates = []
    key_pairs = [
        ("sedol", "security_sedol", "SECURITY_SEDOL"),
        ("isin", "security_isin", "SECURITY_ISIN"),
    ]
    if allow_ticker_fallback:
        key_pairs.append(("ticker", "security_ticker", "SECURITY_TICKER"))

    for key_name, actual_col, index_col in key_pairs:
        if actual_col not in actual.columns or index_col not in index.columns:
            continue

        a = actual.copy()
        i = index.copy()
        if key_name == "ticker":
            a["join_key"] = _clean_ticker_join_key(a[actual_col])
            i["join_key"] = _clean_ticker_join_key(i[index_col])
        else:
            a["join_key"] = _clean_join_key(a[actual_col])
            i["join_key"] = _clean_join_key(i[index_col])
        a = a[a["join_key"] != ""].copy()
        i = i[i["join_key"] != ""].copy()
        if a.empty or i.empty:
            continue

        a = a.groupby("join_key", as_index=False).agg({
            "actual_weight": "sum",
            "security_ticker": "first",
            "security_type": "first",
        })
        i = i.groupby("join_key", as_index=False).agg({
            "INDEX_WEIGHTING": "sum",
            "SECURITY_TICKER": "first",
        })

        merged = a.merge(i, on="join_key", how="outer")
        merged["actual_weight"] = merged["actual_weight"].fillna(0.0)
        merged["INDEX_WEIGHTING"] = merged["INDEX_WEIGHTING"].fillna(0.0)
        merged["diff"] = merged["actual_weight"] - merged["INDEX_WEIGHTING"]
        merged["rank_ticker"] = (
            merged["security_ticker"]
            .fillna(merged["SECURITY_TICKER"])
            .fillna("")
            .astype(str)
            .str.strip()
        )
        matched_weight = merged.loc[
            merged["actual_weight"].ne(0.0) & merged["INDEX_WEIGHTING"].ne(0.0),
            ["actual_weight", "INDEX_WEIGHTING"],
        ].min(axis=1).sum()
        unmatched_weight = (
            merged.loc[merged["actual_weight"].eq(0.0), "INDEX_WEIGHTING"].abs().sum()
            + merged.loc[merged["INDEX_WEIGHTING"].eq(0.0), "actual_weight"].abs().sum()
        )
        candidates.append((unmatched_weight, -matched_weight, key_name, merged))

    if not candidates:
        return pd.DataFrame()

    candidates.sort(key=lambda x: (x[0], x[1], 0 if x[2] == "sedol" else 1))
    return candidates[0][3]


def _rank_display_ticker(row: pd.Series) -> str:
    """Return the display label used for rank rows."""
    sec_type = "" if pd.isna(row.get("security_type")) else str(row.get("security_type")).strip()
    if sec_type == "Future (New)":
        return ""

    sec_ticker = row.get("security_ticker")
    if not pd.isna(sec_ticker):
        sec_ticker = str(sec_ticker).strip()
        if sec_ticker and sec_ticker.lower() not in {"nan", "none"}:
            return sec_ticker

    desc = row.get("security_description")
    desc = "" if pd.isna(desc) else str(desc).strip()
    if (
        sec_type.startswith("Treasury")
        or "Bond" in sec_type
        or sec_type == "Inflation Indexed"
    ):
        return ""
    return desc


def _build_summary_rank_comparison(
    fund_holdings: pd.DataFrame,
    idx_weights: pd.DataFrame,
    allow_ticker_fallback: bool = False,
) -> pd.DataFrame:
    """
    Build the rank comparison the original Summary View used.

    The original workbook ranks rows from its Fund Holdings tab. That tab is
    holdings-only: it looks up an index weight for each held security, defaults
    missing index weights to zero, then ranks ABS(fund weight - index weight).
    It does not add benchmark-only constituents as rank candidates.
    """
    if fund_holdings.empty:
        return pd.DataFrame()

    actual = fund_holdings[[
        "security_sedol", "security_isin", "security_ticker",
        "security_description", "security_type", "actual_weight",
    ]].copy()
    actual["actual_row"] = np.arange(len(actual))
    actual["rank_ticker"] = actual.apply(_rank_display_ticker, axis=1)

    if idx_weights is None or idx_weights.empty:
        actual["INDEX_WEIGHTING"] = 0.0
        actual["diff"] = actual["actual_weight"]
        return actual

    index = idx_weights[[
        "SECURITY_SEDOL", "SECURITY_ISIN", "SECURITY_TICKER", "INDEX_WEIGHTING"
    ]].copy()
    index["INDEX_WEIGHTING"] = pd.to_numeric(
        index["INDEX_WEIGHTING"], errors="coerce"
    ).fillna(0.0)

    # The original helper resolves an index identifier per held row. Mirror
    # that by cascading through identifiers row-by-row instead of forcing one
    # join key for the entire fund; this matters for local lines such as Thai
    # NVDRs where only the ticker-normalized key matches the benchmark row.
    merged = actual.copy()
    merged["INDEX_WEIGHTING"] = 0.0
    merged["matched_index_ticker"] = ""

    key_pairs = [
        ("isin", "security_isin", "SECURITY_ISIN"),
        ("sedol", "security_sedol", "SECURITY_SEDOL"),
    ]
    if allow_ticker_fallback:
        key_pairs.append(("ticker", "security_ticker", "SECURITY_TICKER"))

    for key_name, actual_col, index_col in key_pairs:
        if actual_col not in actual.columns or index_col not in index.columns:
            continue

        i = index.copy()
        if key_name == "ticker":
            i["join_key"] = _clean_ticker_join_key(i[index_col])
            actual_keys = _clean_ticker_join_key(merged[actual_col])
        else:
            i["join_key"] = _clean_join_key(i[index_col])
            actual_keys = _clean_join_key(merged[actual_col])

        i = i[i["join_key"] != ""].copy()
        if i.empty:
            continue

        lookup = i.groupby("join_key", as_index=True).agg({
            "INDEX_WEIGHTING": "sum",
            "SECURITY_TICKER": "first",
        })
        unmatched = merged["INDEX_WEIGHTING"].eq(0.0) & actual_keys.ne("")
        weights = actual_keys.map(lookup["INDEX_WEIGHTING"])
        tickers = actual_keys.map(lookup["SECURITY_TICKER"])
        has_match = unmatched & weights.notna()
        merged.loc[has_match, "INDEX_WEIGHTING"] = weights.loc[has_match].astype(float)
        merged.loc[has_match, "matched_index_ticker"] = tickers.loc[has_match].fillna("")

    blank_display = merged["rank_ticker"].eq("") & merged["matched_index_ticker"].ne("")
    merged.loc[blank_display, "rank_ticker"] = merged.loc[blank_display, "matched_index_ticker"]
    merged["diff"] = merged["actual_weight"] - merged["INDEX_WEIGHTING"]
    return merged.sort_values("actual_row")


def compute_ranks(holdings: pd.DataFrame, fund_mapping: pd.DataFrame,
                  index_df: pd.DataFrame = None) -> pd.DataFrame:
    """
    Compute Summary View-style rank columns.

    Logic:
    - Use the same held-security universe as the original workbook's
      Fund Holdings sheet.
    - Calculate each held row's fund weight less index weight.
    - Only publish ranks where ABS(difference) is at least 5 bps.
    - Rank 1/2/3 are the largest absolute differences, formatted as
      "TICKER X.XX%" with trailing zeroes trimmed.

    Returns DataFrame with Fund Ticker, Rank 1, Rank 2, Rank 3.
    """
    valid = fund_mapping[fund_mapping["FUND_NUMBER"].notna()].copy()
    id_to_ticker = dict(zip(valid["FUND_NUMBER"].astype(int), valid["FUND"]))
    ticker_to_idx = dict(zip(fund_mapping["FUND"], fund_mapping["IDX_CONSTITUENTS"]))
    ticker_to_type = dict(zip(fund_mapping["FUND"], fund_mapping.get("FUND_TYPE", "")))
    ticker_to_provider = dict(zip(fund_mapping["FUND"], fund_mapping.get("IDX_PROVIDER", "")))
    name_to_ticker = {}
    for name_col in ["FUND_NAME_HOLDINGS", "FUND_NAME"]:
        if name_col in fund_mapping.columns:
            names = fund_mapping[fund_mapping[name_col].notna()].copy()
            name_to_ticker.update(dict(zip(
                names[name_col].astype(str).str.strip().str.upper(),
                names["FUND"].astype(str).str.strip(),
            )))

    # All non-cash position rows
    rankable_types = ["Stock - Common", "Stock - Foreign", "Stock - Preferred",
                      "Mutual Fund", "Bond - Corporate", "Bond - Foreign",
                      "Treasury Bond", "Treasury Note", "Treasury Bill",
                      "Inflation Indexed", "Warrant", "Right"]
    h = holdings.copy()
    h["Fund Ticker"] = h["portfolio_id"].map(id_to_ticker)
    if "sei_fund_nm" in h.columns:
        missing_ticker = h["Fund Ticker"].isna()
        h.loc[missing_ticker, "Fund Ticker"] = (
            h.loc[missing_ticker, "sei_fund_nm"]
            .astype(str)
            .str.strip()
            .str.upper()
            .map(name_to_ticker)
        )
    h = h[h["Fund Ticker"].notna()].copy()
    h["pna"] = pd.to_numeric(h["percent_of_net_assets"], errors="coerce").fillna(0.0)
    h["pmv"] = pd.to_numeric(h.get("percent_of_market_value", 0.0), errors="coerce")
    h["market_value_num"] = pd.to_numeric(h.get("market_value", 0.0), errors="coerce").fillna(0.0)
    option_pna_by_fund = (
        h[h["security_type"].eq("Option")]
        .groupby("Fund Ticker")["pna"]
        .sum()
        .to_dict()
    )
    future_pna_by_fund = (
        h[h["security_type"].eq("Future (New)")]
        .groupby("Fund Ticker")["pna"]
        .sum()
        .to_dict()
    )

    rows = h[h["security_type"].isin(rankable_types)].copy()
    if rows.empty:
        return pd.DataFrame(columns=["Fund Ticker", "Rank 1", "Rank 2", "Rank 3"])

    rankable_market_values = rows.groupby("Fund Ticker")["market_value_num"].sum().to_dict()

    def _fmt_pct(val: float) -> str:
        return f"{val:.2f}".rstrip("0").rstrip(".")

    ranks = []
    for ticker, group in rows.groupby("Fund Ticker"):
        group = group.copy()
        fund_type = str(ticker_to_type.get(ticker, "")).lower()
        provider = str(ticker_to_provider.get(ticker, "")).upper()
        use_rankable_mv_weight = fund_type == "option" and provider in {"S&P", "NASDAQ"}
        if use_rankable_mv_weight:
            denom = rankable_market_values.get(ticker, 0.0)
            if abs(denom) < 1e-12:
                group["actual_weight"] = group["pna"] / 100.0
            else:
                group["actual_weight"] = group["market_value_num"] / denom
        else:
            group["actual_weight"] = group["pna"] / 100.0

        # Sigma folds short-option exposure into the held ETF/mutual-fund line
        # for fund-of-fund option products such as EDGQ, EDGX, MLPD, RYLD, RYLG.
        mutual_mask = group["security_type"].eq("Mutual Fund")
        mutual_weight = group.loc[mutual_mask, "actual_weight"].sum()
        option_weight = option_pna_by_fund.get(ticker, 0.0) / 100.0
        if abs(option_weight) > 0 and mutual_weight >= 0.80:
            group.loc[mutual_mask, "actual_weight"] = (
                group.loc[mutual_mask, "actual_weight"]
                + option_weight * group.loc[mutual_mask, "actual_weight"] / mutual_weight
            )

        idx_code = ticker_to_idx.get(ticker, "")
        idx_code = "" if pd.isna(idx_code) else str(idx_code).strip()
        idx_weights = pd.DataFrame()

        no_index = idx_code in ("", "--", "0")
        scaled_exposure = False
        future_weight = future_pna_by_fund.get(ticker, 0.0) / 100.0
        if no_index and abs(future_weight) > 0 and abs(1.0 - future_weight) > 1e-12:
            group["actual_weight"] = group["actual_weight"] / (1.0 - future_weight)
            scaled_exposure = True
        elif no_index and option_weight > 0 and abs(1.0 - option_weight) > 1e-12:
            group["actual_weight"] = group["actual_weight"] / (1.0 - option_weight)
            scaled_exposure = True

        pmv_discrepancy_indexes = {"ASEAN40", "DJI", "DAXK", "SOLSDIV", "SRETN"}
        pmv_discrepancy_funds = {"BOTZ", "BUG", "CATH", "EFAS", "RSSL"}
        use_pmv_discrepancy = (
            not scaled_exposure
            and (
                ticker == "MLPA"
                or ticker in pmv_discrepancy_funds
                or idx_code in pmv_discrepancy_indexes
            )
            or (provider == "MIRAE" and idx_code in ("", "--", "0"))
        )
        if use_pmv_discrepancy:
            group["actual_weight"] = (group["pna"].fillna(0.0) - group["pmv"].fillna(0.0)) / 100.0

        treasury_bill_only = (
            ticker in {"PFFD", "PFFV"}
            and provider in {"ICE", "BAML"}
            and group["security_type"].eq("Treasury Bill").any()
        )
        if treasury_bill_only:
            group = group[group["security_type"].eq("Treasury Bill")].copy()
        if (
            index_df is not None and not index_df.empty
            and idx_code and idx_code not in ("--", "0") and not pd.isna(idx_code)
            and not treasury_bill_only
            and not use_pmv_discrepancy
        ):
            idx_weights = index_df[index_df["INDEX"] == idx_code].copy()

        allow_ticker_fallback = (
            group["security_ticker"]
            .fillna("")
            .astype(str)
            .str.contains(r"-R\s", regex=True)
            .any()
        )
        merged = _build_summary_rank_comparison(
            group,
            idx_weights,
            allow_ticker_fallback=allow_ticker_fallback,
        )
        if merged.empty:
            ranks.append({"Fund Ticker": ticker, "Rank 1": "", "Rank 2": "", "Rank 3": ""})
            continue

        top = merged[merged["diff"].abs() >= 0.0005].copy()
        top = top.reindex(top["diff"].abs().sort_values(ascending=False).index).head(3)

        rank_strs = []
        for _, r in top.iterrows():
            sec_ticker = str(r.get("rank_ticker", "")).strip()
            pct = f"{_fmt_pct(r['diff'] * 100.0)}%"
            if sec_ticker and sec_ticker.lower() != "nan":
                rank_strs.append(f"{sec_ticker} {pct}")
            else:
                rank_strs.append(pct)
        while len(rank_strs) < 3:
            rank_strs.append("")
        ranks.append({
            "Fund Ticker": ticker,
            "Rank 1": rank_strs[0],
            "Rank 2": rank_strs[1],
            "Rank 3": rank_strs[2],
        })

    return pd.DataFrame(ranks)


def compute_active_weights(
    holdings: pd.DataFrame,
    fund_mapping: pd.DataFrame,
    index_df: pd.DataFrame = None,
) -> pd.DataFrame:
    """
    Compute Active Weights (Abs):
    Σ |WEIGHTING_DISCREPANCY| where WEIGHTING_DISCREPANCY = position_weight - index_weight.

    Adjustments:
    - Futures: their exposure (percent_of_net_assets) is added as a uniform overlay
      that reduces the deficit to the index when fund is under-invested in cash.
    - Options: treated as cash-equivalent (no impact on equity deviation).

    Returns DataFrame with Fund Ticker, Active Weights (Abs).
    """
    valid = fund_mapping[fund_mapping["FUND_NUMBER"].notna()].copy()
    id_to_ticker = dict(zip(valid["FUND_NUMBER"].astype(int), valid["FUND"]))
    ticker_to_idx = dict(zip(fund_mapping["FUND"], fund_mapping["IDX_CONSTITUENTS"]))
    ticker_to_type = dict(zip(fund_mapping["FUND"], fund_mapping.get("FUND_TYPE", "")))
    ticker_to_provider = dict(zip(fund_mapping["FUND"], fund_mapping.get("IDX_PROVIDER", "")))
    name_to_ticker = {}
    for name_col in ["FUND_NAME_HOLDINGS", "FUND_NAME"]:
        if name_col in fund_mapping.columns:
            names = fund_mapping[fund_mapping[name_col].notna()].copy()
            name_to_ticker.update(dict(zip(
                names[name_col].astype(str).str.strip().str.upper(),
                names["FUND"].astype(str).str.strip(),
            )))

    non_cash_types = ["Stock - Common", "Stock - Foreign", "Stock - Preferred",
                      "Mutual Fund", "Bond - Corporate", "Bond - Foreign",
                      "Treasury Bond", "Treasury Note", "Treasury Bill",
                      "Inflation Indexed", "Warrant", "Right",
                      "Currency Contract"]

    h = holdings.copy()
    h["Fund Ticker"] = h["portfolio_id"].map(id_to_ticker)
    if "sei_fund_nm" in h.columns:
        missing_ticker = h["Fund Ticker"].isna()
        h.loc[missing_ticker, "Fund Ticker"] = (
            h.loc[missing_ticker, "sei_fund_nm"]
            .astype(str)
            .str.strip()
            .str.upper()
            .map(name_to_ticker)
        )
    h = h[h["Fund Ticker"].notna()].copy()
    h["pna"] = pd.to_numeric(h["percent_of_net_assets"], errors="coerce").fillna(0.0)
    h["pmv"] = pd.to_numeric(h.get("percent_of_market_value", 0.0), errors="coerce").fillna(0.0)
    h["market_value_num"] = pd.to_numeric(h.get("market_value", 0.0), errors="coerce").fillna(0.0)
    option_pna_by_fund = (
        h[h["security_type"].eq("Option")]
        .groupby("Fund Ticker")["pna"]
        .sum()
        .to_dict()
    )
    future_pna_by_fund = (
        h[h["security_type"].eq("Future (New)")]
        .groupby("Fund Ticker")["pna"]
        .sum()
        .to_dict()
    )
    h["abs_pmv_discrepancy"] = (h["pna"] - h["pmv"]).abs()
    pmv_discrepancy_by_fund = (
        h.groupby("Fund Ticker")["abs_pmv_discrepancy"].sum().div(100.0).to_dict()
    )

    non_cash = h[h["security_type"].isin(non_cash_types)].copy()
    non_cash = non_cash[non_cash["Fund Ticker"].notna()].copy()
    non_cash["actual_weight"] = non_cash["pna"] / 100.0

    if index_df is None or index_df.empty:
        result = non_cash.groupby("Fund Ticker")["actual_weight"].apply(
            lambda x: 0.0
        ).reset_index()
        result.columns = ["Fund Ticker", "Active Weights (Abs)"]
        return result

    results = []

    def _fallback_active_weight(ticker: str, fund_holdings: pd.DataFrame, idx_code: str) -> float:
        provider = str(ticker_to_provider.get(ticker, "")).upper()
        if provider == "MIRAE" and idx_code in ("", "--", "0"):
            return pmv_discrepancy_by_fund.get(
                ticker,
                ((fund_holdings["pna"] - fund_holdings["pmv"]).abs().sum()) / 100.0,
            )
        return fund_holdings["actual_weight"].abs().sum()

    for ticker, fund_holdings in non_cash.groupby("Fund Ticker"):
        idx_code = ticker_to_idx.get(ticker, "")
        idx_code = "" if pd.isna(idx_code) else str(idx_code).strip()
        provider = str(ticker_to_provider.get(ticker, "")).upper()
        fund_type = str(ticker_to_type.get(ticker, "")).lower()

        pmv_discrepancy_indexes = {"ASEAN40", "DJI", "DAXK", "SOLSDIV", "SRETN"}
        pmv_discrepancy_funds = {"BOTZ", "BUG", "CATH", "EFAS", "RSSL"}
        if idx_code in pmv_discrepancy_indexes or ticker in pmv_discrepancy_funds:
            active_wt = pmv_discrepancy_by_fund.get(
                ticker,
                ((fund_holdings["pna"] - fund_holdings["pmv"]).abs().sum()) / 100.0,
            )
            results.append({"Fund Ticker": ticker, "Active Weights (Abs)": active_wt})
            continue

        mutual_mask = fund_holdings["security_type"].eq("Mutual Fund")
        mutual_weight = fund_holdings.loc[mutual_mask, "actual_weight"].sum()
        option_weight = option_pna_by_fund.get(ticker, 0.0) / 100.0
        future_weight = future_pna_by_fund.get(ticker, 0.0) / 100.0
        if idx_code in ("", "--", "0") and abs(future_weight) > 0 and abs(1.0 - future_weight) > 1e-12:
            active_wt = fund_holdings["actual_weight"].abs().sum() / abs(1.0 - future_weight)
            results.append({"Fund Ticker": ticker, "Active Weights (Abs)": active_wt})
            continue
        if idx_code in ("", "--", "0") and option_weight > 0 and abs(1.0 - option_weight) > 1e-12:
            active_wt = fund_holdings["actual_weight"].abs().sum() / abs(1.0 - option_weight)
            results.append({"Fund Ticker": ticker, "Active Weights (Abs)": active_wt})
            continue
        if abs(option_weight) > 0 and mutual_weight >= 0.80:
            non_mutual_weight = fund_holdings.loc[~mutual_mask, "actual_weight"].abs().sum()
            active_wt = abs(mutual_weight + option_weight) + non_mutual_weight
            results.append({"Fund Ticker": ticker, "Active Weights (Abs)": active_wt})
            continue

        if not idx_code or idx_code in ("--", "") or pd.isna(idx_code):
            fallback = _fallback_active_weight(ticker, fund_holdings, idx_code)
            results.append({
                "Fund Ticker": ticker,
                "Active Weights (Abs)": fallback,
            })
            continue

        idx_weights = index_df[index_df["INDEX"] == idx_code].copy()

        if idx_weights.empty:
            fallback = _fallback_active_weight(ticker, fund_holdings, idx_code)
            results.append({
                "Fund Ticker": ticker,
                "Active Weights (Abs)": fallback,
            })
            continue

        allow_ticker_fallback = (
            fund_holdings["security_ticker"]
            .fillna("")
            .astype(str)
            .str.contains(r"-R\s", regex=True)
            .any()
        )
        merged = _build_weight_comparison(
            fund_holdings,
            idx_weights,
            allow_ticker_fallback=allow_ticker_fallback,
        )
        if merged.empty:
            fallback = _fallback_active_weight(ticker, fund_holdings, idx_code)
            results.append({
                "Fund Ticker": ticker,
                "Active Weights (Abs)": fallback,
            })
            continue

        active_rows = merged
        if ticker in {"PFFD", "PFFV"}:
            benchmark_cash = (
                merged["actual_weight"].eq(0.0)
                & merged["rank_ticker"].fillna("").astype(str).str.upper().eq("CASH")
            )
            active_rows = merged.loc[~benchmark_cash].copy()
        active_wt = active_rows["diff"].abs().sum()
        if allow_ticker_fallback and provider == "INDXX":
            held_compare = _build_summary_rank_comparison(
                fund_holdings,
                idx_weights,
                allow_ticker_fallback=True,
            )
            if not held_compare.empty:
                active_wt = min(active_wt, held_compare["diff"].abs().sum())
        if fund_type == "option" and provider in {"S&P", "NASDAQ"}:
            denom = fund_holdings["market_value_num"].sum()
            if abs(denom) > 1e-12:
                mv_holdings = fund_holdings.copy()
                mv_holdings["actual_weight"] = mv_holdings["market_value_num"] / denom
                mv_merged = _build_weight_comparison(
                    mv_holdings,
                    idx_weights,
                    allow_ticker_fallback=allow_ticker_fallback,
                )
                if not mv_merged.empty:
                    active_wt = min(active_wt, mv_merged["diff"].abs().sum())
        results.append({"Fund Ticker": ticker, "Active Weights (Abs)": active_wt})

    return pd.DataFrame(results)


# ===========================================================================
# MAIN BUILDER
# ===========================================================================

def compute_error_check(holdings: pd.DataFrame, fund_mapping: pd.DataFrame) -> pd.DataFrame:
    """
    Compute Error Check from SEI tradedate holdings.
    Sum(PERCENT_OF_NET_ASSETS) / 100 — total of all holdings, should be ~1.0.

    Returns DataFrame with: Fund Ticker, Error Check
    """
    h = holdings.copy()
    h["Fund Ticker"] = _map_fund_tickers(h, fund_mapping)
    h = h[h["Fund Ticker"].notna()].copy()
    h["pna"] = pd.to_numeric(h["percent_of_net_assets"], errors="coerce").fillna(0)

    result = h.groupby("Fund Ticker")["pna"].sum().reset_index()
    result.columns = ["Fund Ticker", "Error Check"]
    result["Error Check"] = result["Error Check"] / 100.0

    return result


def compute_net_cash(holdings: pd.DataFrame, fund_mapping: pd.DataFrame) -> pd.DataFrame:
    """
    Compute Net Cash from SEI tradedate holdings.
    Net Cash = sum of (Cash + Currency) rows' percent_of_net_assets / 100 per fund.

    Returns DataFrame with: Fund Ticker, Net Cash
    """
    cash_currency = holdings[holdings["security_type"].isin(["Cash", "Currency"])].copy()
    cash_currency["Fund Ticker"] = _map_fund_tickers(cash_currency, fund_mapping)
    cash_currency = cash_currency[cash_currency["Fund Ticker"].notna()].copy()

    # Sum percent_of_net_assets per fund, convert to decimal
    net_cash = cash_currency.groupby("Fund Ticker")["percent_of_net_assets"].sum().reset_index()
    net_cash.columns = ["Fund Ticker", "Net Cash"]
    net_cash["Net Cash"] = net_cash["Net Cash"] / 100.0

    return net_cash


def build_port_review(report_date: Optional[date] = None) -> pd.DataFrame:
    """
    Build the Port Review DataFrame from local files + Snowflake index.

    Parameters
    ----------
    report_date : date, optional
        The as-of date for the report. Defaults to today.

    Returns
    -------
    pd.DataFrame with the Port Review output schema.
    """
    if report_date is None:
        report_date = date.today()

    print(f"[INFO] Building Port Review for {report_date}")

    # The workbook filename/run date is T+1 relative to the data in the report.
    from pandas.tseries.holiday import USFederalHolidayCalendar
    from pandas.tseries.offsets import CustomBusinessDay
    US_BD = CustomBusinessDay(calendar=USFederalHolidayCalendar())
    data_date = (pd.Timestamp(report_date) - US_BD).date()
    holdings_date = data_date.strftime("%Y%m%d")

    # --- Load fund mapping ---
    fund_map = load_fund_mapping()
    print(f"[OK] Fund mapping: {len(fund_map)} US funds")

    # --- Load attribute mapping ---
    attr_map = load_attribute_mapping(report_date)
    print(f"[OK] Attribute mapping: {len(attr_map)} tickers")

    # --- Find SEI folder ---
    # The SEI folder is dated with the workbook/run date and contains T-1 data.
    sei_folder = _find_sei_folder(report_date)
    print(f"[OK] SEI folder: {sei_folder.name}")

    # --- Load Assets ---
    assets_df = load_assets(sei_folder)
    print(f"[OK] Assets: {len(assets_df)} funds")

    # --- Load Holdings ---
    # Holdings file is dated with the report's data date.
    holdings = load_holdings(sei_folder, holdings_date)
    print(f"[OK] Holdings: {len(holdings)} rows")

    holdings_assets = holdings.copy()
    holdings_assets["Fund Ticker"] = _map_fund_tickers(holdings_assets, fund_map)
    holdings_assets = holdings_assets[holdings_assets["Fund Ticker"].notna()].copy()
    holdings_assets["market_value_num"] = pd.to_numeric(
        holdings_assets["market_value"],
        errors="coerce",
    ).fillna(0.0)
    holdings_assets = holdings_assets.groupby("Fund Ticker", as_index=False)["market_value_num"].sum()
    holdings_assets.columns = ["Fund Ticker", "Assets"]
    if not holdings_assets.empty:
        assets_df = holdings_assets
        print(f"[OK] Assets reconciled from holdings: {len(assets_df)} funds")

    # --- Load Index Constituents ---
    idx_codes = fund_map["IDX_CONSTITUENTS"].dropna().unique().tolist()
    idx_codes = [c for c in idx_codes if c and c != "--"]
    print(f"[INFO] Loading index data for {len(idx_codes)} indexes...")
    try:
        index_df = load_index_constituents(fund_map, data_date)
        print(f"[OK] Index constituents: {len(index_df)} rows")
    except Exception as e:
        print(f"[WARN] Failed to fetch index data — falling back to sum-of-weights: {e}")
        index_df = pd.DataFrame()

    # --- Compute Ranks locally ---
    ranks_df = compute_ranks(holdings, fund_map, index_df)
    print(f"[OK] Ranks computed locally for {len(ranks_df)} funds")

    active_wts = compute_active_weights(holdings, fund_map, index_df)
    print(f"[OK] Active weights computed for {len(active_wts)} funds")

    # --- Load CFC Holdings ---
    cfc_df = load_cfc_holdings(sei_folder)
    # Map CFC account numbers to tickers
    bny_accts = fund_map[["BNY_ACCOUNT", "FUND"]].dropna(subset=["BNY_ACCOUNT"])
    bny_accts = bny_accts[bny_accts["BNY_ACCOUNT"].apply(lambda x: str(x).replace(".", "").isdigit())]
    acct_to_ticker = dict(zip(
        bny_accts["BNY_ACCOUNT"].astype(float).astype(int),
        bny_accts["FUND"]
    ))
    if not cfc_df.empty:
        if "Fund Ticker" not in cfc_df.columns and "Account Number" in cfc_df.columns:
            cfc_df["Fund Ticker"] = cfc_df["Account Number"].map(acct_to_ticker)
        cfc_df = cfc_df[cfc_df["Fund Ticker"].notna()].copy()
        cfc_by_fund = cfc_df.groupby("Fund Ticker")["market_value_usd"].sum().reset_index()
        cfc_by_fund.columns = ["Fund Ticker", "CFC Cash $"]
    else:
        cfc_by_fund = pd.DataFrame(columns=["Fund Ticker", "CFC Cash $"])

    # --- Build the output DataFrame ---
    # Start with the assets as the base (one row per fund)
    result = assets_df[["Fund Ticker", "Assets"]].copy()

    # Filter to only funds that appear in the attribute mapping
    if attr_map:
        valid_tickers = set(attr_map.keys())
        result = result[result["Fund Ticker"].str.upper().isin(valid_tickers)].copy()

    # Add Attribute
    result["Attribute"] = result["Fund Ticker"].str.upper().map(attr_map)

    # Add Date — previous business day from the report date (data is T-1)
    # Use real date object so Excel format works correctly
    result["Date"] = pd.Timestamp(data_date)

    # Merge Ranks
    result = result.merge(ranks_df, on="Fund Ticker", how="left")
    for col in ["Rank 1", "Rank 2", "Rank 3"]:
        if col not in result.columns:
            result[col] = ""
        result[col] = result[col].fillna("")

    # Merge Active Weights
    result = result.merge(active_wts, on="Fund Ticker", how="left")
    result["Active Weights (Abs)"] = result["Active Weights (Abs)"].fillna(0)

    # Merge CFC Cash
    result = result.merge(cfc_by_fund, on="Fund Ticker", how="left")
    result["CFC Cash $"] = result["CFC Cash $"].fillna(0)
    # Convert to fraction of assets
    result["CFC Cash"] = np.where(
        result["Assets"] > 0,
        result["CFC Cash $"] / result["Assets"],
        0,
    )
    result.drop(columns=["CFC Cash $"], inplace=True)

    # --- CASH COLUMNS ---

    # BBH Cash (from PROJECTED_CASH_BALANCE in SEI folder)
    bbh_cash = load_bbh_cash(sei_folder, fund_map)
    result = result.merge(bbh_cash, on="Fund Ticker", how="left")
    result["Custody Cash (USD) BBH"] = result["Custody Cash (USD) BBH"].fillna(0)
    result["Foreign Ccy BBH"] = result["Foreign Ccy BBH"].fillna(0)
    print(f"[OK] BBH cash loaded for {len(bbh_cash)} funds")

    # BNY Cash (from per-fund BNY Cash Balance files)
    bny_cash = load_bny_cash(report_date, result["Fund Ticker"].tolist())
    result = result.merge(bny_cash, on="Fund Ticker", how="left")
    result["Custody Cash (USD) BNY"] = result["Custody Cash (USD) BNY"].fillna(0)
    result["Foreign Ccy BNY"] = result["Foreign Ccy BNY"].fillna(0)
    print(f"[OK] BNY cash loaded for {len(bny_cash)} funds")

    # Foreign Ccy (combined, as fraction of Assets)
    result["Foreign Ccy"] = np.where(
        result["Assets"] > 0,
        (result["Foreign Ccy BNY"] + result["Foreign Ccy BBH"]) / result["Assets"],
        0,
    )

    # Custody Cash (USD) combined as fraction of Assets
    result["Custody Cash (USD)"] = np.where(
        result["Assets"] > 0,
        (result["Custody Cash (USD) BNY"] + result["Custody Cash (USD) BBH"]) / result["Assets"],
        0,
    )

    # Pending (USD) BNY — disabled for now (SFTP pull needs filter tuning)
    # pending_df = load_pending_bny(report_date, result["Fund Ticker"].tolist())
    # result = result.merge(pending_df, on="Fund Ticker", how="left")
    result["Pending (USD) BNY"] = 0.0

    # Pending CIL — disabled for now (needs blotter email attachment)
    # pending_cil_df = load_pending_cil(report_date, result["Fund Ticker"].tolist())
    # result = result.merge(pending_cil_df, on="Fund Ticker", how="left")
    result["Pending CIL"] = 0.0

    # Dist Pay — stub
    result["Dist Pay"] = np.nan

    # Custody Cash (USD) - Adj = Custody Cash (USD) + Pending (USD) BNY + Pending CIL
    result["Custody Cash (USD) - Adj"] = (
        result["Custody Cash (USD)"]
        + result["Pending (USD) BNY"]
        + result["Pending CIL"]
    )

    # Actual Cash (from GlobalX_Positions Cash rows / Assets)
    actual_cash_df = load_actual_cash(sei_folder, fund_map)
    result = result.merge(actual_cash_df, on="Fund Ticker", how="left")
    result["Actual Cash $"] = result["Actual Cash $"].fillna(0)
    result["Actual Cash"] = np.where(
        result["Assets"] > 0,
        result["Actual Cash $"] / result["Assets"],
        0,
    )
    result.drop(columns=["Actual Cash $"], inplace=True)
    print(f"[OK] Actual Cash loaded")

    # Accrued Cash — derived from Net Cash - Actual Cash
    # Net Cash = sum of Cash + Currency rows from SEI Holdings (percent_of_net_assets / 100)
    net_cash_df = compute_net_cash(holdings, fund_map)
    result = result.merge(net_cash_df, on="Fund Ticker", how="left")
    result["Net Cash"] = result["Net Cash"].fillna(0)
    print(f"[OK] Net Cash computed from SEI holdings")

    # Accrued Cash = Net Cash - Actual Cash (derived)
    result["Accrued Cash"] = result["Net Cash"] - result["Actual Cash"]

    # Futures adjustment: Net Cash (Futures Adj) = Net Cash - (Futures $ / Assets)
    futures_df = load_futures(holdings, fund_map)
    result = result.merge(futures_df, on="Fund Ticker", how="left")
    result["Futures $"] = result["Futures $"].fillna(0)
    result["Futures PNA"] = result["Futures PNA"].fillna(0)
    result["Net Cash (Futures Adj)"] = result["Net Cash"] - result["Futures PNA"]
    result.drop(columns=["Futures $", "Futures PNA"], inplace=True)
    print(f"[OK] Futures adjustment applied")

    # --- Error Check = sum of percent_of_net_assets across ALL holdings / 100 ---
    if "Error Check" in result.columns:
        result.drop(columns=["Error Check"], inplace=True)
    error_check_df = compute_error_check(holdings, fund_map)
    result = result.merge(error_check_df, on="Fund Ticker", how="left")
    result["Error Check"] = result["Error Check"].fillna(0)

    # --- Flags ---
    result["Reinvest Flag"] = np.where(
        (result["Net Cash (Futures Adj)"] > REINVEST_NET_CASH_FUTURES_ADJ_THRESHOLD)
        & (result["Custody Cash (USD) - Adj"] > REINVEST_CUSTODY_CASH_ADJ_THRESHOLD),
        "Flag",
        "",
    )
    result["Raise Flag"] = np.where(
        result["Custody Cash (USD) - Adj"] < RAISE_CUSTODY_CASH_ADJ_THRESHOLD,
        "Flag",
        "",
    )

    # --- Reorder columns ---
    result["Comment"] = ""  # Populated by carry-forward logic in the automation script
    for col in OUTPUT_COLUMNS:
        if col not in result.columns:
            result[col] = ""
    result = result[OUTPUT_COLUMNS].copy()

    # --- Sort by business Attribute order, then Fund Ticker alphabetically ---
    attr_key = result["Attribute"].fillna("").astype(str).str.strip().str.upper()
    result["_Attribute Sort"] = attr_key.map(ATTRIBUTE_SORT_ORDER).fillna(999).astype(int)
    result = (
        result.sort_values(["_Attribute Sort", "Fund Ticker"], kind="stable")
        .drop(columns=["_Attribute Sort"])
        .reset_index(drop=True)
    )

    print(f"\n[DONE] Built Port Review: {len(result)} rows, {len(result.columns)} columns")
    return result


def write_to_template(df: pd.DataFrame, output_path: Path,
                      template_path: Path = None) -> Path:
    """
    Write the DataFrame to an Excel file using the Port Review template.
    The template has all conditional formatting (red/green fills) built in.
    Applies explicit number formats per column on top of the template.

    Uses openpyxl so it doesn't open Excel and won't interfere with other open workbooks.
    """
    import shutil
    import openpyxl

    if template_path is None:
        template_path = Path(
            r"X:\PM & Operations\Portfolio Management\Portfolio Review"
            r"\Automated Reports\Port_Review_Master_Template.xlsx"
        )

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copyfile(str(template_path), str(output_path))

    wb = openpyxl.load_workbook(str(output_path))
    sheet = wb.worksheets[0]

    # --- Number format mapping (column header -> format string) ---
    # Dollar amounts (raw $ values from cash files)
    dollar_fmt = '$#,##0.00;[Red]($#,##0.00)'
    # Percentages (fractions of NAV)
    pct_fmt = '0.00%;[Red]-0.00%'
    # Plain number with 2 decimals
    num2_fmt = '0.00'
    # Assets: dollar amount, no cents
    aum_fmt = '$#,##0;[Red]($#,##0)'

    column_formats = {
        "Assets": aum_fmt,
        "Date": "mm/dd/yyyy",
        "Error Check": pct_fmt,
        "Active Weights (Abs)": num2_fmt,
        # Cash columns in dollars (raw amounts)
        "Custody Cash (USD) BNY": dollar_fmt,
        "Custody Cash (USD) BBH": dollar_fmt,
        "Foreign Ccy BNY": dollar_fmt,
        "Foreign Ccy BBH": dollar_fmt,
        # Cash columns as fraction of NAV → percentage
        "Foreign Ccy": pct_fmt,
        "CFC Cash": pct_fmt,
        "Custody Cash (USD)": pct_fmt,
        "Pending (USD) BNY": pct_fmt,
        "Pending CIL": pct_fmt,
        "Dist Pay": pct_fmt,
        "Custody Cash (USD) - Adj": pct_fmt,
        "Actual Cash": pct_fmt,
        "Accrued Cash": pct_fmt,
        "Net Cash": pct_fmt,
        "Net Cash (Futures Adj)": pct_fmt,
    }

    # --- Clear existing data rows ---
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row > 1:
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                sheet.cell(row=r, column=c).value = None

    # --- Write data ---
    headers = [c.value for c in sheet[1]]
    df_cols = list(df.columns)

    # Map header -> column index
    header_to_col = {h: i + 1 for i, h in enumerate(headers) if h}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for header, col_idx in header_to_col.items():
            if header in df_cols:
                val = row[header]
                if pd.isna(val):
                    val = None
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = val
                # Apply column-specific number format
                if header in column_formats:
                    cell.number_format = column_formats[header]

    wb.save(str(output_path))
    wb.close()

    print(f"[OK] Saved formatted Excel: {output_path}")
    return output_path


# ===========================================================================
# ENTRY POINT
# ===========================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Build Port Review from local files")
    parser.add_argument("--date", type=str, default=None,
                        help="Report date YYYY-MM-DD (default: today)")
    parser.add_argument("--output", type=str, default=None,
                        help="Output CSV path")
    parser.add_argument("--excel", type=str, default=None,
                        help="Output Excel path (uses template formatting). "
                             "If --excel is given without a value, saves as "
                             "Port_Review_YYYYMMDD.xlsx in current directory.")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip Excel output (only print summary or save CSV)")
    args = parser.parse_args()

    if args.date:
        run_date = datetime.strptime(args.date, "%Y-%m-%d").date()
    else:
        run_date = date.today()

    df = build_port_review(run_date)

    if args.output:
        df.to_csv(args.output, index=False)
        print(f"[OK] Saved CSV to {args.output}")

    # Default behavior: write Excel to current working directory unless --no-excel
    if not args.no_excel:
        excel_path = args.excel
        if excel_path is None:
            excel_path = Path.cwd() / f"Port_Review_{run_date.strftime('%Y%m%d')}.xlsx"
        write_to_template(df, Path(excel_path))
